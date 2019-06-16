import openpyxl
from Bio import SeqIO


doc=openpyxl.load_workbook('GENE.xlsx') #Abrir documento 

#### LECTURA DE HOJAS:
hoja = doc.get_sheet_by_name('Hoja1') # Leer la hoja de Excel que se llama 'Hoja1'

nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column # Número de columnas.


lista=[]
for r in range(7495,11689): # En python se empieza en 1 + la cabecera (1) = 2.
    ID = hoja.cell(r,1).value
    lista.append(ID)
    

for ids in lista:
    fasta_sequences = SeqIO.parse(open('../TARA.faa'),'fasta')
    for fasta in fasta_sequences:
        name, sequence = fasta.id, str(fasta.seq)
        if fasta.id == ids:
            c=0
            for amino in range(0,len(sequence)-4):
                if sequence[amino]=='T' and sequence[amino+1] == 'V' and sequence[amino+2]=='P' and sequence[amino+3]=='L':
                    c=1
                    if sequence[amino+4] == 'L':
                        aa='Leu'
                        print(name,aa,'Green')
                    elif sequence[amino+4] == 'Q':
                        aa='Gln'
                        print(name,aa,'Blue')
                    elif sequence[amino+4] == 'M':
                        aa='Met'
                        print(name,aa,'Green')
                    else:
                        print(name, sequence[amino+4])
                    
                elif c == 0 and amino == (len(sequence)-5):
                    print(name,'ND','ND')
    fasta_sequences.close()

            
