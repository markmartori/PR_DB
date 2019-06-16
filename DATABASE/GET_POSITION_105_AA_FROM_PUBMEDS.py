import openpyxl
from Bio import SeqIO


doc=openpyxl.load_workbook('GENE_PUBMED.xlsx') #Abrir documento 

#### LECTURA DE HOJAS:
hoja = doc.get_sheet_by_name('Hoja1') # Leer la hoja de Excel que se llama 'Hoja1'

nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column # Número de columnas.


lista=[]
for r in range(2,nrows+1): # En python se empieza en 1 + la cabecera (1) = 2.
    ID = hoja.cell(r,1).value
    lista.append(ID)
    

for ids in lista:
    fasta_sequences = SeqIO.parse(open('allpro.faa'),'fasta')
    for fasta in fasta_sequences:
        name, sequence = fasta.id, str(fasta.seq)
        if fasta.id == ids:
            c=0
            for amino in range(0,len(sequence)-4):
                if sequence[amino]=='T' and sequence[amino+1] == 'V' and sequence[amino+2]=='P' and sequence[amino+3]=='L':
                    c=1
                    print(name,sequence[amino+4])
                    
                elif c == 0 and amino == (len(sequence)-5):
                    print(name,'ND')
    fasta_sequences.close()

            
