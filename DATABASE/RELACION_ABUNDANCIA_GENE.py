##################################
########### UTILITY ##############
##################################

'''
THIS FILE READS AN EXCEL DOCUMENT.

READS THE FILE AND CREATES A DICTIONARY WITH:
- GENE NAME AS KEY.
- NAME OF THE SAMPLES WHERE THIS GENE HAS BEEN FOUND.

THERE ARE 179 SAMPLES AND ALL GENES HAS BEEN FOUND IN MORE THAN 10 OF THEM.

IT PRINTS THE DICTIONARY INTO A NEW FILE FOLLOWING:
-Gene1 1stSamplethegeneappears.
-Gene1 2ndSamplethegeneappears.
...
...
'''

import openpyxl
import collections


doc=openpyxl.load_workbook('EXCEL.xlsx') # Load archivo

hoja = doc.get_sheet_by_name('hoja1') # Coger la primera hoja.

filas=hoja.rows # Todas las filas en tupla.
nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column

'''
dsample={} # Diccionario con SAMPLE y Coordenada.

for a in range(2,ncols):
    sample=hoja.cell(1,a).value
    dsample[sample]=hoja.cell(1,a).coordinate

#print(dsample)
'''

dicc={}
for a in range(2,nrows+1):
    for b in range(2,ncols+1):
        val = hoja.cell(a,b).value
        if val != 0:
            samp=hoja.cell(row=1,column=b).value
            gene=hoja.cell(row=a,column=1).value
            if gene in dicc.keys():
                dicc[gene].append(samp)
            else:
                dicc[gene]=[samp]
#print(dicc)


archivo = open("GENE_SAMPLE_PRUEBA.txt", "w")

for keys in dicc.keys():
    for val in dicc[keys]:
        archivo.write(str(keys)+str(val)+'\n')
        
archivo.close() 


