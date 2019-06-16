import openpyxl

doc=openpyxl.load_workbook('EXCEL.xlsx')

#### LECTURA DE HOJAS:

nombre_hojas<-doc.get_sheet_names() # Nombre de todas las hojas.

hoja = doc.get_sheet_by_name('Sheet1') # Hoja que se llama 'Sheet1'

hoja.title # Nombre de la hoja.

#### LECTURA DE CELDAS:

# FILAS:
filas = hoja.rows # Todas las filas en TUPLA.

for rows in filas:
    print(rows) # Todas las filas vistas mejor organizadas para ir fila por fila.

# CELDAS:
hoja['A1'].value # Valor A1 de nuestra hoja.
# output: u'Last' 

hoja.cell(row=1,column=1).value # Valor row1 col1 de nuestra hoja.

hoja['F2'].value
#output: 33040

type(hoja['F2'].value)
#output: <type 'int'>

for row in filas: # Recorrer todos los valores de la hoja, no es la forma más eficiente.
    for col in row:
        print(col.value,)
    print("")
    
ultimafila = hoja.get_highest_row() # Valor de la última fila.
ultimacol=hoja.get_highest_column() # Valor de la última col.


from ex.cell import get_column_letter

get_column_letter(hoja.get_highest_column()) # Letra de la última columna.

#### RECORRIDO DE FILAS Y COLUMNAS EFICIENTE:

for row in filas:
    for col in row:
        print(columna.coordinate,columna.value)
    print("----Final de Fila----")


#### SELECCION RECTANGULAR O DE AREA:

seleccion = hoja['A1':'F5']
for row in seleccion:
    for col in row:
        print(col.coordinate,col.value)
    print("----Final de Fila----")


#####################                                       ##################
##################### CONECTAR PYTHON CON LA BASE DE DATOS ###################
####################                                        #################
    
database_host = localhost
username = root
password = PRicm_2019
database_name = mark

import pymysql
# Abre conexion con la BD:
db = pymysql.connect("db_host","username","password","db")

# Preparar cursor object:
cursor = db.cursor()

# Ejecuta el SQL query usando el metodo execute():
cursor.execute("SELECT VERSION()")

# Procesa una única linea usando el metodo fetchone():
data = cursor.fetchone()
print ("Database version : {0}".format(data)) #Print

# Desconecta del servidor:
db.close()

                     
##### INSERTAR FILAS EN LA BASE DE DATOS:

# Abre conexion con la BD:
db = pymysql.connect("db_host","username","password","db")

# prepare a cursor object using cursor() method
cursor = db.cursor()

# Prepare SQL query to INSERT a record into the database.
sql = "INSERT INTO test(id, name, email) \
   VALUES (NULL,'{0}','{1}')".format("cosme","testmail@sever.com")
try:
   # Execute the SQL command
   cursor.execute(sql)
   # Commit your changes in the database
   db.commit()
except:
   # Rollback in case there is any error
   db.rollback()


# desconectar del servidor
db.close()

##### LEER BASE DE DATOS:

# Open database connection
db = pymysql.connect("db_host","username","password","db")


# prepare a cursor object using cursor() method
cursor = db.cursor()

# Prepare SQL query to READ a record into the database.
sql = "SELECT * FROM test \
WHERE id > {0}".format(0)

# Execute the SQL command
cursor.execute(sql)

# Fetch all the rows in a list of lists.
results = cursor.fetchall()
for row in results:
   id = row[0]
   name = row[1]
   email = row[2]
   # Now print fetched result
   print ("id = {0}, name = {1}, email = {1}".format(id,name,email))


# disconnect from server
db.close()












