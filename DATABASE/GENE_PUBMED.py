################################################
################# UTILITY ######################
################################################
'''
THIS FILE READS AN EXCEL DOCUMENT and connects to the DataBase with:
host
user
Password
db name
charset

IT QUERYS AN INSERT INTO GENE_PUBMED ENTITY. CREATING THE COLUMN 'GeneID' AND 22 MORE. THIS PROGRAM IS INSERTING THE METADATA COLLECTED FROM ALL THE PUBMED PUBLICATIONS STUDIED.

'''

import pymysql.cursors
import openpyxl


doc=openpyxl.load_workbook('EXCEL.xlsx')

#### LECTURA DE HOJAS:
hoja = doc.get_sheet_by_name('Hoja1') # Hoja que se llama Hoja1'

filas = hoja.rows # Todas las filas en TUPLA.
nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column # Número de columnas.



# Connect to the database
connection = pymysql.connect(host='host',
                             user='username',
                             password='password',
                             db='name',charset='',
                             cursorclass=pymysql.cursors.DictCursor)

cursor = connection.cursor()
# Create a new record
sql = "INSERT INTO `GENE_PUBMED` (`GeneID`,`PUBMEDID`,`CULTURED`,`LATITUDE`,`LONGITUDE`,`WATER_POSITION`,`DEPTH`,`GEOGRAPHY`,`MONTH`,`TEMPERATURE`,`MEDIUM`,`LOW_FILTER`,`UPPER_FILTER`,`EXTRACTION_METHOD`,`LENGTH_NUC`,`ABSORPTION`,`pH`,`CHLOROPHYLL_A`,`GENE_TRANSFER`,`SEASON`,`PROVINCE`,`NUM_PROVINCE`,`Ocean`) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

# Query INSERT

for r in range(2,nrows+1): # En python se empieza en 1 + la cabecera (1) = 2.
    GeneID= hoja.cell(r,1).value
    PUBMEDID= hoja.cell(r,2).value
    CULTURED= hoja.cell(r,3).value
    LATITUDE= hoja.cell(r,4).value
    LONGITUDE= hoja.cell(r,5).value
    WATER_POSITION= hoja.cell(r,6).value
    DEPTH= hoja.cell(r,7).value
    GEOGRAPHY= hoja.cell(r,8).value
    MONTH= hoja.cell(r,9).value
    TEMPERATURE=hoja.cell(r,10).value
    MEDIUM=hoja.cell(r,11).value
    LOW_FILTER= hoja.cell(r,12).value
    UPPER_FILTER= hoja.cell(r,13).value
    EXTRACTION_METHOD= hoja.cell(r,14).value
    LENGTH_NUC= hoja.cell(r,15).value
    ABSORPTION= hoja.cell(r,16).value
    pH= hoja.cell(r,17).value
    CHLOROPHYLL_A= hoja.cell(r,18).value
    GENE_TRANSFER= hoja.cell(r,19).value
    SEASON=hoja.cell(r,20).value
    PROVINCE=hoja.cell(r,21).value
    NUM_PROVINCE=hoja.cell(r,22).value
    Ocean=hoja.cell(r,23).value
    
    # Asignar valores:
    values = (GeneID,PUBMEDID,CULTURED,LATITUDE,LONGITUDE,WATER_POSITION,DEPTH,GEOGRAPHY,MONTH,TEMPERATURE,MEDIUM,LOW_FILTER,UPPER_FILTER,EXTRACTION_METHOD,LENGTH_NUC,ABSORPTION,pH,CHLOROPHYLL_A,GENE_TRANSFER,SEASON,PROVINCE,NUM_PROVINCE,Ocean)

    # Ejecutar sql Query
    cursor.execute(sql,values)     # Ejecutar COUNT


    # Salir del Loop.
# Cerrar el cursor:
cursor.close()

# Commit.
connection.commit()

# Cerrar base de datos:
connection.close()

# Print Results:
print("")
print("All Done!")
print("")
columns = str(ncols)
rows = str(nrows)
print("I just imported " + columns +  " column(s) and " + rows + " rows to MySQL!")
