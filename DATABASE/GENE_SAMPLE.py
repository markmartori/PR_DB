################################
########## UTILITY #############
################################
'''
THIS FILE READS AN EXCEL DOCUMENT and connects to the DataBase with:
host
user
Password
db name
charset

IT QUERYS AN INSERT INTO GENE_SAMPLE ENTITY. CREATING THE COLUMN 'GeneID' AND SAMPLEIDE.

'''


import pymysql.cursors
import openpyxl


doc=openpyxl.load_workbook('EXCEL.xlsx')

#### LECTURA DE HOJAS:
hoja = doc.get_sheet_by_name('hoja1') # Hoja que se llama Hoja1'

filas = hoja.rows # Todas las filas en TUPLA.
nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column # Número de columnas.


# Connect to the database
connection = pymysql.connect(host='host',
                             user='username',
                             password='password',
                             db='name',charset='',
                             cursorclass=pymysql.cursors.DictCursor)

cursor = connection.cursor()
# Create a new record
sql = "INSERT INTO `GENE_SAMPLE` (`GeneID`,`SAMPLEID`) VALUES (%s,%s)"     # Query INSERT


for r in range(2,nrows+1): # En python se empieza en 1 + la cabecera (1) = 2.
    GeneID= hoja.cell(r,1).value
    SAMPLEID = hoja.cell(r,2).value

    
    # Asignar valores:
    values = (GeneID,SAMPLEID)

    # Ejecutar sql Query
    cursor.execute(sql,values)     # Ejecutar COUNT
    connection.commit()

    # Salir del Loop.
# Cerrar el cursor:
cursor.close()

# Commit.
connection.commit()

# Cerrar base de datos:
connection.close()

# Print Results:

print("")
print("All Done!")
print("")
columns = str(ncols)
rows = str(nrows)
print("I just imported " + columns +  " column(s) and " + rows + " rows to MySQL!")

