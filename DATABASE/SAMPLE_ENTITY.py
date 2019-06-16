################################
########## UTILITY #############
################################
'''
THIS FILE READS AN EXCEL DOCUMENT and connects to the DataBase with:
host
user
Password
db name
charset

IT QUERYS AN INSERT INTO SAMPLE ENTITY. CREATING THE COLUMN 'SAMPLE' AND 119 MORE.

THIS PROGRAM INSERTS ALL THE DATA COLLECTED IN TARA OCEANS.

'''

import pymysql.cursors
import openpyxl


doc=openpyxl.load_workbook('SAMPLE.xlsx')

#### LECTURA DE HOJAS:
hoja = doc.get_sheet_by_name('environmental_parameters') # Hoja que se llama Hoja1'

filas = hoja.rows # Todas las filas en TUPLA.
nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column # Número de columnas.



# Connect to the database
connection = pymysql.connect(host='host',
                             user='username',
                             password='password',
                             db='name',charset='',
                             cursorclass=pymysql.cursors.DictCursor)

cursor = connection.cursor()
# Create a new record
sql = "INSERT INTO `SAMPLE` (`SAMPLEID`,`STATION`,`DEPTH_ZONE`,`LATITUDE`,`LONGITUDE`,`OGA_ID`,`ENA_ID`,`INSDC_ID`,`BioSamples_ID`,`Sample_method`,`Depth_Nominal`,`Environmental_feature`,`LOW_FILTER`,`UPPER_FILTER`,`Biome`,`Region`,`Province`,`Barcode`,`Date_Time`,`Seafloor`,`Temperature`,`Salinity`,`Density`,`Distance_coast`,`Chlorophyll_A`,`Depth`,`PAR`,`O2`,`NO3`,`Iron_5m`,`Ammonium_5m`,`Nitrite_5m`,`Nitrate_5m`,`CDOM`,`NPP_C`,`POC`,`PIC`,`Alkalinity`,`Carbon_Total`,`CO2`,`CO3`,`HCO3`,`pH`,`NO2`,`PO4`,`NO3_NO2`,`Si`,`MLD`,`DCM`,`Depth_max_B_V_freq`,`Depth_max_O2`,`Depth_min_O2`,`Depth_Nitracline`,`Shannon_Darwin`,`Shannon_Physat`,`miTAG_SILVIA_Chao`,`miTAG_SILVIA_Shannon`,`OG_Richness`,`OG_Evenness`,`Chlorophyll_c3`,`Peridinin`,`Fucoxanthin`,`Prasinoxanthin`,`Hexanoyloxyfucoxanthin`,`Alloxanthin`,`Zeexanthin`,`Lutein`,`Sea_ice`,`Season`,`Season_subperiod`,`Okubo_Weiss`,`Lyapunov`,`Residence_time`,`MONTH`,`Code_MP`,`Leg`,`Pres`,`Layer`,`LayerNum`,`WaterMassType`,`a254`,`SR`,`MaxZ`,`LongHurstProvince`,`Conductivity`,`Fluo`,`SPAR`,`Turb_FTU`,`SalinityWOA`,`SiO4`,`ProvinceNum`,`Sigma`,`O2_sat`,`AOU_corr_umol_kg`,`Fmax1_resp_prok`,`Fmax2_resp_euk`,`Fmax3_tirosina`,`Fmax4_triptofano`,`TEP`,`POC_uM`,`pmol_leu`,`SE`,`LNA`,`HNA`,`All_BT`,`percentHNA`,`cell_size`,`Bacterial_cell_C`,`Biomass`,`ugC_l_d`,`d_1`,`turnover_days`,`HNF`,`low_virus`,`medium_virus`,`high_virus`,`all_virus`,`VBR`,`Ocean`)VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

for r in range(2,nrows+1): # En python se empieza en 1 + la cabecera (1) = 2.

    SAMPLEID= hoja.cell(r,1).value
    STATION= hoja.cell(r,2).value
    DEPTH_ZONE= hoja.cell(r,3).value
    LATITUDE= hoja.cell(r,4).value
    LONGITUDE= hoja.cell(r,5).value
    OGA_ID= hoja.cell(r,6).value
    ENA_ID= hoja.cell(r,7).value
    INSDC_ID= hoja.cell(r,8).value
    BioSamples_ID= hoja.cell(r,9).value
    Sample_method= hoja.cell(r,10).value
    Depth_Nominal= hoja.cell(r,11).value
    Environmental_feature= hoja.cell(r,12).value
    LOW_FILTER= hoja.cell(r,13).value
    UPPER_FILTER= hoja.cell(r,14).value
    Biome= hoja.cell(r,15).value
    Region= hoja.cell(r,16).value
    Province= hoja.cell(r,17).value
    Barcode= hoja.cell(r,18).value
    Date_Time= hoja.cell(r,19).value
    Seafloor= hoja.cell(r,20).value
    Temperature= hoja.cell(r,21).value
    Salinity= hoja.cell(r,22).value
    Density= hoja.cell(r,23).value
    Distance_coast= hoja.cell(r,24).value
    Chlorophyll_A= hoja.cell(r,25).value
    Depth= hoja.cell(r,26).value
    PAR= hoja.cell(r,27).value
    O2= hoja.cell(r,28).value
    NO3= hoja.cell(r,29).value
    Iron_5m= hoja.cell(r,30).value
    Ammonium_5m= hoja.cell(r,31).value
    Nitrite_5m= hoja.cell(r,32).value
    Nitrate_5m= hoja.cell(r,33).value
    CDOM= hoja.cell(r,34).value
    NPP_C= hoja.cell(r,35).value
    POC= hoja.cell(r,36).value
    PIC= hoja.cell(r,37).value
    Alkalinity= hoja.cell(r,38).value
    Carbon_Total= hoja.cell(r,39).value
    CO2= hoja.cell(r,40).value
    CO3= hoja.cell(r,41).value
    HCO3= hoja.cell(r,42).value
    pH= hoja.cell(r,43).value
    NO2= hoja.cell(r,44).value
    PO4= hoja.cell(r,45).value
    NO3_NO2= hoja.cell(r,46).value
    Si= hoja.cell(r,47).value
    MLD= hoja.cell(r,48).value
    DCM= hoja.cell(r,49).value
    Depth_max_B_V_freq= hoja.cell(r,50).value
    Depth_max_O2= hoja.cell(r,51).value
    Depth_min_O2= hoja.cell(r,52).value
    Depth_Nitracline= hoja.cell(r,53).value
    Shannon_Darwin= hoja.cell(r,54).value
    Shannon_Physat= hoja.cell(r,55).value
    miTAG_SILVIA_Chao= hoja.cell(r,56).value
    miTAG_SILVIA_Shannon= hoja.cell(r,57).value
    OG_Richness= hoja.cell(r,58).value
    OG_Evenness= hoja.cell(r,59).value
    Chlorophyll_c3= hoja.cell(r,60).value
    Peridinin= hoja.cell(r,61).value
    Fucoxanthin= hoja.cell(r,62).value
    Prasinoxanthin= hoja.cell(r,63).value
    Hexanoyloxyfucoxanthin= hoja.cell(r,64).value
    Alloxanthin= hoja.cell(r,65).value
    Zeexanthin= hoja.cell(r,66).value
    Lutein= hoja.cell(r,67).value
    Sea_ice= hoja.cell(r,68).value
    Season= hoja.cell(r,69).value
    Season_subperiod= hoja.cell(r,70).value
    Okubo_Weiss= hoja.cell(r,71).value
    Lyapunov= hoja.cell(r,72).value
    Residence_time= hoja.cell(r,73).value
    MONTH= hoja.cell(r,74).value
    Code_MP= hoja.cell(r,75).value
    Leg= hoja.cell(r,76).value
    Pres= hoja.cell(r,77).value
    Layer= hoja.cell(r,78).value
    LayerNum= hoja.cell(r,79).value
    WaterMassType= hoja.cell(r,80).value
    a254= hoja.cell(r,81).value
    SR= hoja.cell(r,82).value
    MaxZ= hoja.cell(r,83).value
    LongHurstProvince= hoja.cell(r,84).value
    Conductivity= hoja.cell(r,85).value
    Fluo= hoja.cell(r,86).value
    SPAR= hoja.cell(r,87).value
    Turb_FTU= hoja.cell(r,88).value
    SalinityWOA= hoja.cell(r,89).value
    SiO4= hoja.cell(r,90).value
    ProvinceNum= hoja.cell(r,91).value
    Sigma= hoja.cell(r,92).value
    O2_sat= hoja.cell(r,93).value
    AOU_corr_umol_kg= hoja.cell(r,94).value
    Fmax1_resp_prok= hoja.cell(r,95).value
    Fmax2_resp_euk= hoja.cell(r,96).value
    Fmax3_tirosina= hoja.cell(r,97).value
    Fmax4_triptofano= hoja.cell(r,98).value
    TEP= hoja.cell(r,99).value
    POC_uM= hoja.cell(r,100).value
    pmol_leu= hoja.cell(r,101).value
    SE= hoja.cell(r,102).value
    LNA= hoja.cell(r,103).value
    HNA= hoja.cell(r,104).value
    All_BT= hoja.cell(r,105).value
    percentHNA= hoja.cell(r,106).value
    cell_size= hoja.cell(r,107).value
    Bacterial_cell_C= hoja.cell(r,108).value
    Biomass= hoja.cell(r,109).value
    ugC_l_d= hoja.cell(r,110).value
    d_1= hoja.cell(r,111).value
    turnover_days= hoja.cell(r,112).value
    HNF= hoja.cell(r,113).value
    low_virus= hoja.cell(r,114).value
    medium_virus= hoja.cell(r,115).value
    high_virus= hoja.cell(r,116).value
    all_virus= hoja.cell(r,117).value
    VBR= hoja.cell(r,118).value
    Ocean= hoja.cell(r,119).value

    
    # Asignar valores:
    values = (SAMPLEID,STATION,DEPTH_ZONE,LATITUDE,LONGITUDE,OGA_ID,ENA_ID,INSDC_ID,BioSamples_ID,Sample_method,Depth_Nominal,Environmental_feature,LOW_FILTER,UPPER_FILTER,Biome,Region,Province,Barcode,Date_Time,Seafloor,Temperature,Salinity,Density,Distance_coast,Chlorophyll_A,Depth,PAR,O2,NO3,Iron_5m,Ammonium_5m,Nitrite_5m,Nitrate_5m,CDOM,NPP_C,POC,PIC,Alkalinity,Carbon_Total,CO2,CO3,HCO3,pH,NO2,PO4,NO3_NO2,Si,MLD,DCM,Depth_max_B_V_freq,Depth_max_O2,Depth_min_O2,Depth_Nitracline,Shannon_Darwin,Shannon_Physat,miTAG_SILVIA_Chao,miTAG_SILVIA_Shannon,OG_Richness,OG_Evenness,Chlorophyll_c3,Peridinin,Fucoxanthin,Prasinoxanthin,Hexanoyloxyfucoxanthin,Alloxanthin,Zeexanthin,Lutein,Sea_ice,Season,Season_subperiod,Okubo_Weiss,Lyapunov,Residence_time,MONTH,Code_MP,Leg,Pres,Layer,LayerNum,WaterMassType,a254,SR,MaxZ,LongHurstProvince,Conductivity,Fluo,SPAR,Turb_FTU,SalinityWOA,SiO4,ProvinceNum,Sigma,O2_sat,AOU_corr_umol_kg,Fmax1_resp_prok,Fmax2_resp_euk,Fmax3_tirosina,Fmax4_triptofano,TEP,POC_uM,pmol_leu,SE,LNA,HNA,All_BT,percentHNA,cell_size,Bacterial_cell_C,Biomass,ugC_l_d,d_1,turnover_days,HNF,low_virus,medium_virus,high_virus,all_virus,VBR
,Ocean)
    

    # Ejecutar sql Query
    cursor.execute(sql,values)

    # Salir del Loop.
# Cerrar el cursor:
cursor.close()

# Commit.
connection.commit()

# Cerrar base de datos:
connection.close()

# Print Results:

print("")
print("All Done!")
print("")
columns = str(ncols)
rows = str(nrows)
print("I just imported " + columns + " columns and " + rows + " rows to MySQL!")
