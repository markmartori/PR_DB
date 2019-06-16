
import pymysql.cursors
import openpyxl


doc=openpyxl.load_workbook('GENE_SAMPLE_MALASPINA.xlsx')
hoja = doc.get_sheet_by_name('hoja1') # Coger la primera hoja.
ncols=hoja.max_column

for a in range(1,2):
    for b in range(1,ncols):
        v=hoja.cell(row=1,column=b).value
        print(v)
    
