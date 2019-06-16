################################
########## UTILITY #############
################################
'''
THIS FILE READS AN EXCEL DOCUMENT and connects to the DataBase with:
host
user
Password
db name
charset

IT QUERYS AN INSERT INTO PUBMEDID ENTITY. CREATING THE COLUMN 'PUBMEDID'.

'''


import pymysql.cursors
import openpyxl


doc=openpyxl.load_workbook('PUBMED.xlsx')

#### LECTURA DE HOJAS:
hoja = doc.get_sheet_by_name('Hoja1') # Hoja que se llama Hoja1'

filas = hoja.rows # Todas las filas en TUPLA.
nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column # Número de columnas.


# Connect to the database
connection = pymysql.connect(host='host',
                             user='username',
                             password='password',
                             db='name',charset='',
                             cursorclass=pymysql.cursors.DictCursor)

cursor = connection.cursor()
# Create a new record
sqlINSERT = "INSERT INTO `PUBMED` (`PubmedID`) VALUES (%s)"     # Query INSERT
sqlCOUNT = "SELECT COUNT(*) FROM PUBMED WHERE PubmedID = '%s'"  # Hacemos un COUNT para evitar duplicaciones.
# Contador de registros Únicos.
nunique=0
for r in range(2,nrows): # En python se empieza en 1 + la cabecera (1) = 2.
    PubmedID= hoja.cell(r,1).value

    
    # Asignar valores:
    values = (PubmedID)

    # Ejecutar sql Query
    cursor.execute(sqlCOUNT,values)     # Ejecutar COUNT
    (number_of_rows)=cursor.fetchone()  # Devuelve un DICCIONARIO.
    
    if number_of_rows['COUNT(*)'] != 1: # Si el valor de la key del DICC. no es 1
        nunique +=1
        cursor.execute(sqlINSERT,values)


    # Salir del Loop.
# Cerrar el cursor:
cursor.close()

# Commit.
connection.commit()

# Cerrar base de datos:
connection.close()

# Print Results:

print("")
print("All Done!")
print("")
columns = str(ncols)
rows = str(nrows)
print("I just imported " + columns +  " column(s) and " + str(nunique)+ " out of " + rows + " rows to MySQL!")
