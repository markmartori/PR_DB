##################################
########### UTILITY ##############
##################################

'''
THIS FILE READS AN EXCELL DOCUMENT NAMED "GENE_SAMPLE_MALASPINA.xlsx".

READS THE FILE AND CREATES A DICTIONARY WITH:
- GENE NAME AS KEY.
- NAME OF THE SAMPLES WHERE THIS GENE HAS BEEN FOUND.

THERE ARE 116 SAMPLES AND ALL GENES HAS BEEN FOUND IN MORE THAN 10 OF THEM.

IT PRINTS THE DICTIONARY INTO A NEW FILE FOLLOWING:
-Gene1 1stSamplethegeneappears.
-Gene1 2ndSamplethegeneappears.

'''

import openpyxl
import collections


doc=openpyxl.load_workbook('EXCEL.xlsx') # Load archivo

hoja = doc.get_sheet_by_name('hoja1') # Coger la primera hoja.

filas=hoja.rows # Todas las filas en tupla.
nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column



dicc={}
for a in range(2,nrows+1):
    for b in range(2,ncols+1):
        val = hoja.cell(a,b).value
        if val != 0:
            samp=hoja.cell(row=1,column=b).value
            gene=hoja.cell(row=a,column=1).value
            if gene in dicc.keys():
                dicc[gene].append(samp)
            else:
                dicc[gene]=[samp]


archivo = open("File.xlsx", "w")

for keys in dicc.keys():
    for val in dicc[keys]:
        archivo.write(str(keys)+ '\t' + str(val)+'\n')
        
archivo.close() 


