import openpyxl
from Bio import SeqIO


doc=openpyxl.load_workbook('SAMPLE.xlsx') #Abrir documento 'GENE_PUBMED.xlsx'

#### LECTURA DE HOJAS:
hoja = doc.get_sheet_by_name('environmental_parameters') # Leer la hoja de Excel que se llama 'Hoja1'

nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column # Número de columnas.


lista=[]
for r in range(2,nrows+1): # En python se empieza en 1 + la cabecera (1) = 2.
    ID = hoja.cell(r,26).value
    if ID:
        a=round(float(ID),5)
        print(a)
    else:
        print('')
