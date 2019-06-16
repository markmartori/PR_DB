################################################
################# UTILITY ######################
################################################
'''
THIS FILE READS AN EXCEL DOCUMENT and connects to the DataBase with:
host
user
Password
db name
charset

IT QUERYS AN INSERT INTO GENE ENTITY. CREATING THE COLUMN 'GENEID' AND INSERTING AS MANY RECORDS
AS ROWS HAS THE EXCEL FILE.
'''




import pymysql.cursors
import openpyxl


doc=openpyxl.load_workbook('EXCEL.xlsx')

#### LECTURA DE HOJAS:
hoja = doc.get_sheet_by_name('Hoja1') # Hoja que se llama Hoja1'

filas = hoja.rows # Todas las filas en TUPLA.
nrows=hoja.max_row # Número de filas.
ncols=hoja.max_column # Número de columnas.



# Connect to the database
connection = pymysql.connect(host='host',
                             user='username',
                             password='password',
                             db='db_name',charset='',
                             cursorclass=pymysql.cursors.DictCursor)

cursor = connection.cursor()
# Create a new record
sql = "INSERT INTO `GENE` (`GeneID`,`GeneProvidence`,`PubmedID`,`TAXID`,`Sequencefasta`,`RhodopsinType`,`Domain`,`Kingdom`,`Phylum`,`Class`,`Order`,`Family`,`Genus`,`Species`,`Pos105`,`Color_Tuning`) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

for r in range(2,nrows+1): # En python se empieza en 1 + la cabecera (1) = 2.
    GeneID= hoja.cell(r,1).value
    GeneProvidence= hoja.cell(r,2).value
    PubmedID= hoja.cell(r,3).value
    TAXID= hoja.cell(r,4).value
    Sequencefasta= hoja.cell(r,5).value
    RhodopsinType= hoja.cell(r,6).value
    Domain= hoja.cell(r,7).value
    Kingdom= hoja.cell(r,8).value
    Phylum= hoja.cell(r,9).value
    Class= hoja.cell(r,10).value
    Order= hoja.cell(r,11).value
    Family= hoja.cell(r,12).value
    Genus= hoja.cell(r,13).value
    Species= hoja.cell(r,14).value
    Pos105= hoja.cell(r,15).value
    Color_Tuning= hoja.cell(r,16).value
    
    
    # Asignar valores:
    values = (GeneID,GeneProvidence,PubmedID,TAXID,Sequencefasta,RhodopsinType,Domain,Kingdom,Phylum,Class,Order,Family,Genus,Species,Pos105,Color_Tuning)

    # Ejecutar sql Query
    cursor.execute(sql,values)

    # Salir del Loop.
# Cerrar el cursor:
cursor.close()

# Commit.
connection.commit()

# Cerrar base de datos:
connection.close()

# Print Results:

print("")
print("All Done!")
print("")
columns = str(ncols)
rows = str(nrows)
print("I just imported " + columns + " columns and " + rows + " rows to MySQL!")
